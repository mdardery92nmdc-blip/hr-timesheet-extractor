import os
import sys
import io
import re
import tempfile
from pathlib import Path
from datetime import datetime
import calendar

import streamlit as st
import pandas as pd
import numpy as np

try:
    import pdfplumber
    PDF_PLUMBER_AVAILABLE = True
except ImportError:
    PDF_PLUMBER_AVAILABLE = False

try:
    import easyocr
    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import analysis module
import attendance_analysis as att_an

# ----------------------------------------------------------------------
# App config
# ----------------------------------------------------------------------
st.set_page_config(
    page_title="HR Timesheet Extractor Pro", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header { font-size: 2.5rem; font-weight: bold; color: #1f77b4; }
    .sub-header { font-size: 1.2rem; color: #666; margin-bottom: 2rem; }
    .metric-card { background: #f0f2f6; padding: 1rem; border-radius: 0.5rem; }
    .stAlert { margin: 1rem 0; }
    .report-table { font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-header">📊 HR Timesheet Extractor Pro</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Extract timesheets from PDFs → Generate Comp-Off & Leave Reports</p>', unsafe_allow_html=True)

# ----------------------------------------------------------------------
# OCR engines (cached)
# ----------------------------------------------------------------------
@st.cache_resource(show_spinner="Loading OCR engine...")
def load_easyocr():
    return easyocr.Reader(['en'], gpu=False) if EASYOCR_AVAILABLE else None

ocr_reader = load_easyocr()

# ----------------------------------------------------------------------
# Extraction functions
# ----------------------------------------------------------------------
def extract_employee_info(pdf_path, filename=""):
    employee_info = {
        "name": "Not Found", 
        "number": "Not Found",
        "designation": "Not Found", 
        "company": "Not Found"
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            if not text and ocr_reader:
                try:
                    img = pdf.pages[0].to_image(resolution=200).original
                    ocr_results = ocr_reader.readtext(np.array(img), detail=0)
                    text = "\n".join(ocr_results)
                    st.info(f"ℹ️ [{filename}] OCR fallback used for employee info.")
                except Exception as e:
                    st.warning(f"⚠️ [{filename}] OCR failed: {e}")

            if text:
                STOP = r'(?=\s*(?:NUMBER|DESIGNATION|COMPANY|EMPLOYEE|\n|$))'
                name = re.search(r'NAME\s*[:\s]+([^:\n]+?)' + STOP, text, re.IGNORECASE)
                num  = re.search(r'NUMBER\s*[:\s]+(\d+)', text, re.IGNORECASE)
                des  = re.search(r'DESIGNATION\s*[:\s]+([^:\n]+?)' + STOP, text, re.IGNORECASE)
                com  = re.search(r'COMPANY\s*[:\s]+([^:\n]+?)' + STOP, text, re.IGNORECASE)
                if name: employee_info["name"] = name.group(1).strip()
                if num:  employee_info["number"] = num.group(1).strip()
                if des:  employee_info["designation"] = des.group(1).strip()
                if com:  employee_info["company"] = com.group(1).strip()
    except Exception as e:
        st.warning(f"⚠️ [{filename}] extraction error: {e}")
    return employee_info

def extract_attendance_codes(pdf_path, filename=""):
    attendance = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text_check = pdf.pages[0].extract_text()
            if not text_check and ocr_reader:
                try:
                    img = pdf.pages[0].to_image(resolution=200).original
                    ocr_text = "\n".join(ocr_reader.readtext(np.array(img), detail=0))
                    st.info(f"ℹ️ [{filename}] OCR used for attendance.")
                except Exception:
                    pass

            tables = pdf.pages[0].extract_tables()
            if tables:
                for table in tables:
                    df_tab = pd.DataFrame(table)
                    for idx, row in df_tab.iterrows():
                        row_str = " ".join([str(c).upper() if c else "" for c in row])
                        if "ATTENDANCE" in row_str:
                            # Try header mapping
                            if idx > 0:
                                header_row = df_tab.iloc[idx-1]
                                header_map = {}
                                for col_idx, val in enumerate(header_row):
                                    try:
                                        d = int(str(val).strip())
                                        if 1 <= d <= 31:
                                            header_map[d] = col_idx
                                    except:
                                        pass
                                if header_map:
                                    for d, col_idx in header_map.items():
                                        val = row.iloc[col_idx]
                                        attendance[d] = str(val).strip() if val else ""
                                    return attendance
                            # Fallback: positional columns 1..31
                            for col_idx, val in enumerate(row):
                                if 1 <= col_idx <= 31:
                                    attendance[col_idx] = str(val).strip() if val else ""
                            return attendance
    except Exception as e:
        st.warning(f"⚠️ [{filename}] attendance error: {e}")
    return attendance

# ----------------------------------------------------------------------
# Excel formatter with weekday row
# ----------------------------------------------------------------------
def create_formatted_excel(df_results, day_weekday_map):
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_results.to_excel(writer, index=False, sheet_name="Timesheets", startrow=2)
        worksheet = writer.sheets["Timesheets"]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, col_name in enumerate(df_results.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_num, col_name in enumerate(df_results.columns, 1):
            if col_name.startswith("Day "):
                day_num = int(col_name.split()[1])
                wd = day_weekday_map.get(day_num, "")
                cell = worksheet.cell(row=2, column=col_num)
                cell.value = wd
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(italic=True, size=10)

        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

        worksheet.freeze_panes = "G3"
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

# ----------------------------------------------------------------------
# Display Reports Function (missing from original)
# ----------------------------------------------------------------------
def display_reports(comp_report, leave_report):
    """Display comp-off and leave reports in Streamlit."""

    st.subheader("📋 Comp-Off Report")
    if not comp_report.empty:
        st.success(f"Found {len(comp_report)} employees with comp-off earnings")
        st.dataframe(
            comp_report.sort_values("Comp-Off Earned (Days)", ascending=False),
            use_container_width=True,
            hide_index=True
        )

        # Download comp-off report
        csv_comp = comp_report.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download Comp-Off Report (CSV)",
            data=csv_comp,
            file_name=f"CompOff_Report_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="comp_off_dl"
        )
    else:
        st.info("No comp-off earned this month.")

    st.divider()

    st.subheader("📋 Leave Report")
    if not leave_report.empty:
        st.dataframe(
            leave_report.sort_values("Total Leave Days", ascending=False),
            use_container_width=True,
            hide_index=True
        )

        # Download leave report
        csv_leave = leave_report.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download Leave Report (CSV)",
            data=csv_leave,
            file_name=f"Leave_Report_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="leave_dl"
        )
    else:
        st.info("No leave records found.")

# ----------------------------------------------------------------------
# Sidebar Configuration
# ----------------------------------------------------------------------
with st.sidebar:
    st.header("⚙️ Configuration")

    st.subheader("📅 Period Selection")
    selected_month = st.selectbox(
        "Month", 
        range(1, 13),
        format_func=lambda x: datetime(2000, x, 1).strftime('%B'),
        index=datetime.now().month - 1
    )
    selected_year = st.number_input(
        "Year", 
        min_value=2000, 
        max_value=2100,
        value=datetime.now().year, 
        step=1
    )

    st.divider()

    st.subheader("📝 Leave Codes")
    leave_codes_input = st.text_input(
        "Leave codes (comma separated)", 
        value="L, S, U",
        help="Codes that represent leave days"
    )
    leave_codes = [c.strip().upper() for c in leave_codes_input.split(",")]

    st.divider()

    st.info("""
    **How to use:**
    1. Upload Contract file (CSV/Excel)
    2. Upload Timesheet PDFs
    3. Click Extract & Analyze
    4. Download reports
    """)

    st.divider()
    st.caption("v2.0 | HR Timesheet Extractor Pro")

# ----------------------------------------------------------------------
# Main UI
# ----------------------------------------------------------------------
st.header("📁 Data Upload")

col1, col2 = st.columns(2)

with col1:
    contract_file = st.file_uploader(
        "📄 Employee Contracts (Excel/CSV)",
        type=['xlsx', 'csv'],
        help="Must have columns: 'Employee #' and 'Contractual Days Per Week'"
    )

with col2:
    uploaded_files = st.file_uploader(
        "📤 Timesheet PDFs", 
        type="pdf", 
        accept_multiple_files=True,
        help="Upload one or more timesheet PDFs"
    )

# Validation
if not PDF_PLUMBER_AVAILABLE:
    st.error("❌ pdfplumber not installed. PDF extraction will not work.")
if not OPENPYXL_AVAILABLE:
    st.warning("⚠️ openpyxl not installed. Excel downloads will not work.")

# Process button
if uploaded_files and contract_file:
    st.divider()
    if st.button(f"🚀 Extract & Analyze {len(uploaded_files)} File(s)", use_container_width=True, type="primary"):

        with st.spinner("Processing..."):
            # Load contracts
            try:
                if contract_file.name.endswith('.csv'):
                    contracts_df = pd.read_csv(contract_file)
                else:
                    contracts_df = pd.read_excel(contract_file)

                if 'Employee #' not in contracts_df.columns or 'Contractual Days Per Week' not in contracts_df.columns:
                    st.error("❌ Contract file must have 'Employee #' and 'Contractual Days Per Week' columns.")
                    st.stop()

                contracts_df['Employee #'] = contracts_df['Employee #'].astype(str)
                st.success(f"✅ Loaded {len(contracts_df)} employee contracts")

            except Exception as e:
                st.error(f"❌ Error loading contract file: {e}")
                st.stop()

            # Process each PDF
            all_results = []
            progress_bar = st.progress(0)

            for file_idx, uploaded_file in enumerate(uploaded_files):
                progress = (file_idx + 1) / len(uploaded_files)
                progress_bar.progress(progress, text=f"Processing {uploaded_file.name}...")

                tmp = None
                try:
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                        tmp.write(uploaded_file.getbuffer())
                        tmp_path = Path(tmp.name)

                    emp_info = extract_employee_info(tmp_path, uploaded_file.name)
                    att_codes = extract_attendance_codes(tmp_path, uploaded_file.name)

                except Exception as e:
                    st.error(f"❌ Error processing {uploaded_file.name}: {e}")
                    continue
                finally:
                    if tmp is not None:
                        Path(tmp.name).unlink(missing_ok=True)

                row = {
                    "Serial #": file_idx + 1,
                    "Employee #": emp_info["number"],
                    "Employee Name": emp_info["name"],
                    "Designation": emp_info["designation"],
                    "Company": emp_info["company"]
                }
                for d in range(1, 32):
                    row[f"Day {d}"] = att_codes.get(d, "")
                all_results.append(row)

            progress_bar.empty()

            if not all_results:
                st.error("❌ No data extracted from PDFs.")
                st.stop()

            df_results = pd.DataFrame(all_results)

            # Build day -> weekday map
            day_weekday = {}
            for d in range(1, 32):
                try:
                    date = datetime(selected_year, selected_month, d)
                    day_weekday[d] = date.strftime("%A")
                except ValueError:
                    day_weekday[d] = ""

            # Show summary
            st.divider()
            st.header("📈 Extraction Summary")

            met1, met2, met3, met4 = st.columns(4)
            met1.metric("Total Employees", len(df_results))
            met2.metric("Files Processed", len(uploaded_files))
            met3.metric("Data Format", "Wide (Horizontal)")

            # Count missing contracts
            emp_ids = df_results["Employee #"].astype(str).apply(att_an.normalize_id)
            contract_ids = contracts_df["Employee #"].apply(att_an.normalize_id)
            missing = set(emp_ids) - set(contract_ids)
            met4.metric("Missing Contracts", len(missing), delta="⚠️" if missing else "✅")

            if missing:
                st.warning(f"⚠️ Missing contracts for: {', '.join(missing)}")

            st.divider()
            st.subheader("📋 Extracted Data Preview")
            st.dataframe(df_results, use_container_width=True, hide_index=True)

            # ---------- Comp-Off & Leave Reports ----------
            st.divider()
            st.header("📊 Attendance Analysis Reports")

            try:
                comp_report, leave_report = att_an.calculate_comp_off_and_leave(
                    df_wide=df_results,
                    contracts_df=contracts_df,
                    month=selected_month,
                    year=selected_year,
                    leave_codes=leave_codes
                )
                display_reports(comp_report, leave_report)
            except Exception as e:
                st.error(f"❌ Error generating reports: {e}")
                st.exception(e)

            # ---------- Download Wide Excel ----------
            st.divider()
            st.header("💾 Export Data")

            col_dl1, col_dl2 = st.columns(2)

            if OPENPYXL_AVAILABLE:
                with col_dl1:
                    excel_data = create_formatted_excel(df_results, day_weekday)
                    st.download_button(
                        label="📊 Download Wide Excel (with weekdays)",
                        data=excel_data,
                        file_name=f"Timesheet_Wide_{selected_year}{selected_month:02d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.info("Install openpyxl for Excel export")

            with col_dl2:
                # Raw CSV download
                csv_data = df_results.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📄 Download Raw Data (CSV)",
                    data=csv_data,
                    file_name=f"Timesheet_Raw_{selected_year}{selected_month:02d}.csv",
                    mime="text/csv",
                    use_container_width=True
                )

else:
    st.info("👆 Please upload both the contract file and timesheet PDFs to begin.")

    # Show sample data info
    with st.expander("📖 Sample Contract File Format"):
        sample_df = pd.DataFrame({
            "Employee #": ["005000", "005001"],
            "Contractual Days Per Week": [5.5, 6.0]
        })
        st.dataframe(sample_df, hide_index=True)
        st.code("Employee #,Contractual Days Per Week\n005000,5.5\n005001,6")
